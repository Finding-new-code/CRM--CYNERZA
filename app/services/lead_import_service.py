"""
Service layer for Lead Import operations.
Handles file parsing, field mapping, validation, duplicate detection, and import execution.
"""
import csv
import io
import re
from typing import List, Dict, Any, Tuple, Optional
from sqlalchemy.orm import Session
from fastapi import HTTPException, status

from app.models.lead import Lead, LeadSource
from app.models.import_session import ImportSession, ImportSessionStatus
from app.crud import lead_import as import_crud
from app.crud import lead as lead_crud
from app.schemas.lead import LeadCreate
from app.schemas.lead_import import (
    ValidationError,
    NormalizedLead,
    DuplicateMatch,
    SmartMatch,
    ExistingLeadInfo,
    ImportDataInfo,
    InFileDuplicate
)


# CRM fields available for mapping
AVAILABLE_CRM_FIELDS = ["full_name", "email", "phone", "source"]

# Common column name variations for auto-suggestion
COLUMN_MAPPINGS = {
    "full_name": ["full_name", "fullname", "name", "full name", "contact name", "lead name"],
    "email": ["email", "email address", "e-mail", "mail"],
    "phone": ["phone", "phone number", "telephone", "mobile", "cell", "contact number"],
    "source": ["source", "lead source", "origin", "channel"]
}


def parse_csv_file(file_content: bytes, filename: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Parse CSV file and return columns and data.
    
    Args:
        file_content: Raw file bytes
        filename: Original filename
        
    Returns:
        Tuple of (column_names, list of row dicts)
    """
    try:
        # Try to decode as UTF-8, fallback to latin-1
        try:
            content = file_content.decode('utf-8')
        except UnicodeDecodeError:
            content = file_content.decode('latin-1')
        
        reader = csv.DictReader(io.StringIO(content))
        columns = reader.fieldnames or []
        rows = list(reader)
        
        return list(columns), rows
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse CSV file: {str(e)}"
        )


def parse_excel_file(file_content: bytes, filename: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Parse Excel (.xlsx) file and return columns and data.
    
    Args:
        file_content: Raw file bytes
        filename: Original filename
        
    Returns:
        Tuple of (column_names, list of row dicts)
    """
    try:
        from openpyxl import load_workbook
        
        # Load workbook from bytes
        workbook = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
        sheet = workbook.active
        
        if not sheet:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file has no active sheet"
            )
        
        rows_iter = sheet.iter_rows(values_only=True)
        
        # First row is headers
        header_row = next(rows_iter, None)
        if not header_row:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file has no header row"
            )
        
        # Clean up column names (remove None values, convert to string)
        columns = [str(col).strip() if col is not None else f"Column_{i}" for i, col in enumerate(header_row)]
        
        # Parse data rows
        rows = []
        for row in rows_iter:
            # Skip completely empty rows
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(columns):
                    # Convert cell value to string, handle None
                    row_dict[columns[i]] = str(cell).strip() if cell is not None else ""
            rows.append(row_dict)
        
        workbook.close()
        return columns, rows
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse Excel file: {str(e)}"
        )


def parse_file(file_content: bytes, filename: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Parse file based on extension (CSV or Excel).
    
    Args:
        file_content: Raw file bytes
        filename: Original filename
        
    Returns:
        Tuple of (column_names, list of row dicts)
    """
    filename_lower = filename.lower()
    
    if filename_lower.endswith('.csv'):
        return parse_csv_file(file_content, filename)
    elif filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
        return parse_excel_file(file_content, filename)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file format. Please upload a CSV or Excel (.xlsx) file."
        )


def suggest_mappings(columns: List[str]) -> Dict[str, str]:
    """
    Suggest column to CRM field mappings based on column names.
    
    Args:
        columns: List of detected column names
        
    Returns:
        Dict mapping column names to suggested CRM fields
    """
    suggestions = {}
    
    for column in columns:
        col_lower = column.lower().strip()
        
        for crm_field, variations in COLUMN_MAPPINGS.items():
            if col_lower in variations:
                suggestions[column] = crm_field
                break
    
    return suggestions


def validate_email(email: str) -> bool:
    """Simple email validation."""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email.strip())) if email else False


def normalize_row(row: Dict[str, Any], mappings: Dict[str, str]) -> Dict[str, Any]:
    """
    Normalize a row based on column mappings.
    
    Args:
        row: Raw row data
        mappings: Column to CRM field mappings
        
    Returns:
        Normalized lead data
    """
    normalized = {}
    
    for csv_col, crm_field in mappings.items():
        if csv_col in row:
            value = row[csv_col]
            if value:
                normalized[crm_field] = str(value).strip()
    
    return normalized


def validate_normalized_data(
    data: List[Dict[str, Any]]
) -> Tuple[List[NormalizedLead], List[ValidationError], int]:
    """
    Validate normalized lead data.
    
    Args:
        data: List of normalized lead dicts
        
    Returns:
        Tuple of (valid_leads, validation_errors, valid_count)
    """
    valid_leads = []
    errors = []
    valid_count = 0
    
    for idx, row in enumerate(data):
        row_errors = []
        
        # Check required fields
        if not row.get("full_name"):
            row_errors.append(ValidationError(
                row=idx + 1,
                field="full_name",
                error="Full name is required"
            ))
        
        email = row.get("email", "")
        if not email:
            row_errors.append(ValidationError(
                row=idx + 1,
                field="email",
                error="Email is required"
            ))
        elif not validate_email(email):
            row_errors.append(ValidationError(
                row=idx + 1,
                field="email",
                error="Invalid email format"
            ))
        
        if row_errors:
            errors.extend(row_errors)
        else:
            valid_count += 1
            valid_leads.append(NormalizedLead(
                full_name=row.get("full_name", ""),
                email=row.get("email", ""),
                phone=row.get("phone"),
                source=row.get("source")
            ))
    
    return valid_leads, errors, valid_count


def detect_duplicates(
    db: Session,
    normalized_data: List[Dict[str, Any]]
) -> Tuple[List[DuplicateMatch], List[SmartMatch], List[InFileDuplicate]]:
    """
    Detect duplicates against existing leads and within the import file.
    
    Args:
        db: Database session
        normalized_data: List of normalized lead data
        
    Returns:
        Tuple of (exact_duplicates, smart_matches, in_file_duplicates)
    """
    exact_duplicates = []
    smart_matches = []
    in_file_duplicates = []
    
    # Get all existing lead emails for duplicate checking
    existing_leads = db.query(Lead.id, Lead.full_name, Lead.email).all()
    existing_emails = {lead.email.lower(): lead for lead in existing_leads}
    
    # Track emails within import file
    import_emails: Dict[str, List[int]] = {}
    
    for idx, row in enumerate(normalized_data):
        email = row.get("email", "").lower()
        
        if not email:
            continue
        
        # Check against existing leads
        if email in existing_emails:
            existing = existing_emails[email]
            exact_duplicates.append(DuplicateMatch(
                import_row=idx + 1,
                existing_lead=ExistingLeadInfo(
                    id=existing.id,
                    full_name=existing.full_name,
                    email=existing.email
                ),
                import_data=ImportDataInfo(
                    full_name=row.get("full_name", ""),
                    email=row.get("email", "")
                )
            ))
        
        # Track for in-file duplicates
        if email in import_emails:
            import_emails[email].append(idx + 1)
        else:
            import_emails[email] = [idx + 1]
    
    # Build in-file duplicates list
    for email, rows in import_emails.items():
        if len(rows) > 1:
            in_file_duplicates.append(InFileDuplicate(rows=rows))
    
    return exact_duplicates, smart_matches, in_file_duplicates


def execute_import(
    db: Session,
    session: ImportSession,
    duplicate_decisions: Dict[str, str],
    user_id: int
) -> Tuple[int, int, int, int]:
    """
    Execute the final import.
    
    Args:
        db: Database session
        session: Import session
        duplicate_decisions: Map of row index to action
        user_id: User performing the import
        
    Returns:
        Tuple of (imported_count, skipped_count, updated_count, error_count)
    """
    imported = 0
    skipped = 0
    updated = 0
    errors = 0
    
    file_data = session.file_data or []
    mappings = session.column_mappings or {}
    
    # Normalize all data
    normalized_data = [normalize_row(row, mappings) for row in file_data]
    
    # Get existing leads for duplicate handling
    existing_leads = db.query(Lead).all()
    existing_by_email = {lead.email.lower(): lead for lead in existing_leads}
    
    for idx, data in enumerate(normalized_data):
        row_key = str(idx + 1)
        email = data.get("email", "").lower()
        
        if not email or not data.get("full_name"):
            errors += 1
            continue
        
        # Check if this is a duplicate
        is_duplicate = email in existing_by_email
        action = duplicate_decisions.get(row_key, "skip" if is_duplicate else "create")
        
        try:
            if action == "skip":
                skipped += 1
            elif action == "update" and is_duplicate:
                # Update existing lead
                existing = existing_by_email[email]
                existing.full_name = data.get("full_name", existing.full_name)
                if data.get("phone"):
                    existing.phone = data["phone"]
                db.commit()
                updated += 1
            elif action == "create" or not is_duplicate:
                # Create new lead
                source = LeadSource.OTHER
                if data.get("source"):
                    source_str = data["source"].upper()
                    if hasattr(LeadSource, source_str):
                        source = LeadSource[source_str]
                
                lead_data = LeadCreate(
                    full_name=data["full_name"],
                    email=data["email"],
                    phone=data.get("phone"),
                    source=source
                )
                lead_crud.create_lead(db, lead_data, created_by_id=user_id)
                imported += 1
        except Exception:
            errors += 1
            db.rollback()
    
    return imported, skipped, updated, errors
